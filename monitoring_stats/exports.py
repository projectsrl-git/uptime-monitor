from datetime import timedelta

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side,
)

from django.db.models import Avg
from django.utils import timezone

from check.models import Check

from .services import (
    PERIODS,
    get_monitor_statistics,
)

# ======================================================
# COLORI
# ======================================================

HEADER_FILL = "343A40"
HEADER_FONT = "FFFFFF"

TITLE_FILL = "212529"
TITLE_FONT = "FFFFFF"

SUBTITLE_FILL = "495057"
SUBTITLE_FONT = "FFFFFF"

TOTAL_FILL = "E9ECEF"
TOTAL_FONT = "212529"

BORDER_COLOR = "CED4DA"

UP_FILL = "D1E7DD"
UP_FONT = "0F5132"

DOWN_FILL = "F8D7DA"
DOWN_FONT = "842029"


# ======================================================
# STILI
# ======================================================

THIN_BORDER = Border(
    bottom=Side(
        style="thin",
        color=BORDER_COLOR,
    )
)


def style_header(sheet, row):
    for cell in sheet[row]:

        cell.fill = PatternFill(
            fill_type="solid",
            fgColor=HEADER_FILL,
        )

        cell.font = Font(
            bold=True,
            color=HEADER_FONT,
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        cell.border = THIN_BORDER


def style_total_row(sheet, row, start_column, end_column):
    for column in range(
        start_column,
        end_column + 1,
    ):

        cell = sheet.cell(
            row=row,
            column=column,
        )

        cell.fill = PatternFill(
            fill_type="solid",
            fgColor=TOTAL_FILL,
        )

        cell.font = Font(
            bold=True,
            color=TOTAL_FONT,
        )

        cell.border = THIN_BORDER


def format_datetime(value):
    if value is None:
        return None

    if timezone.is_aware(value):
        value = timezone.localtime(value)

    return value.replace(tzinfo=None)


def format_duration(seconds):
    if seconds is None:
        return None

    return timedelta(seconds=int(seconds))


def make_sheet_name(monitor):
    name = str(monitor.name).strip()

    sheet_name = f"M{monitor.id} - {name}"

    return sheet_name[:31]


def make_unique_sheet_name(
    workbook,
    monitor,
):
    base = make_sheet_name(monitor)

    name = base
    counter = 2

    while name in workbook.sheetnames:

        suffix = f" ({counter})"

        name = f"{base[:31 - len(suffix)]}" f"{suffix}"

        counter += 1

    return name


def style_status_cell(cell, status):
    status = str(status or "").lower()

    if status == "up":

        cell.fill = PatternFill(
            fill_type="solid",
            fgColor=UP_FILL,
        )

        cell.font = Font(
            bold=True,
            color=UP_FONT,
        )

    elif status == "down":

        cell.fill = PatternFill(
            fill_type="solid",
            fgColor=DOWN_FILL,
        )

        cell.font = Font(
            bold=True,
            color=DOWN_FONT,
        )


# ======================================================
# WORKBOOK
# ======================================================


def build_statistics_workbook(
    monitors,
    period,
    include_summary=True,
):
    workbook = Workbook()

    # Rimuove il foglio vuoto creato automaticamente.
    default_sheet = workbook.active

    if default_sheet is not None:
        workbook.remove(default_sheet)

    now = timezone.now()

    # --------------------------------------------------
    # CALCOLIAMO LE STATISTICHE UNA SOLA VOLTA
    # --------------------------------------------------

    monitor_data = []

    for monitor in monitors:

        statistics = get_monitor_statistics(
            monitor,
            period,
            now,
        )

        monitor_data.append(
            {
                "monitor": monitor,
                "statistics": statistics,
            }
        )

    # ==================================================
    # RIEPILOGO
    # ==================================================

    if include_summary:

        summary = workbook.create_sheet("Riepilogo")

        # Titolo
        summary.merge_cells("A1:I1")

        summary["A1"] = "ESPORTA DATI"

        summary["A1"].fill = PatternFill(
            fill_type="solid",
            fgColor=TITLE_FILL,
        )

        summary["A1"].font = Font(
            bold=True,
            size=16,
            color=TITLE_FONT,
        )

        summary["A1"].alignment = Alignment(
            vertical="center",
        )

        summary.row_dimensions[1].height = 28

        # Informazioni export
        summary.merge_cells("A2:I2")

        summary["A2"] = (
            f"Periodo: {period}   |   "
            f"Generato il: "
            f"{timezone.localtime(now):%d/%m/%Y %H:%M}"
        )

        summary["A2"].font = Font(
            color="495057",
            italic=True,
        )

        summary.row_dimensions[2].height = 20

        # Spazio
        summary.row_dimensions[3].height = 8

        # Header
        headers = [
            "Monitor",
            "Status",
            "Uptime",
            "Response medio (ms)",
            "Checks",
            "Checks OK",
            "Checks KO",
            "Incidenti",
            "Downtime",
        ]

        for column, header in enumerate(
            headers,
            start=1,
        ):
            summary.cell(
                row=4,
                column=column,
            ).value = header

        style_header(
            summary,
            4,
        )

        summary.row_dimensions[4].height = 24

        # --------------------------------------------------
        # DATI
        # --------------------------------------------------

        uptime_values = []

        checks_total = 0
        successful_total = 0
        failed_total = 0
        incidents_total = 0
        downtime_total = 0

        selected_monitor_ids = [item["monitor"].id for item in monitor_data]

        period_start = now - PERIODS[period]

        selected_checks = Check.objects.filter(
            monitor_id__in=selected_monitor_ids,
            executed_at__gte=period_start,
            executed_at__lte=now,
        )

        selected_response_average = selected_checks.aggregate(
            average=Avg("response_time_ms")
        )["average"]

        if selected_response_average is not None:
            selected_response_average = round(
                selected_response_average,
                2,
            )

        summary_row = 5

        for item in monitor_data:

            monitor = item["monitor"]

            statistics = item["statistics"]

            summary_data = statistics["summary"]

            response_data = statistics["response_time"]

            uptime = summary_data["uptime_percentage"]

            if uptime is not None:
                uptime_values.append(uptime)

            checks_total += summary_data["checks"]

            successful_total += summary_data["successful_checks"]

            failed_total += summary_data["failed_checks"]

            incidents_total += summary_data["incidents"]

            downtime_total += summary_data["downtime_seconds"]

            row = summary_row

            summary.append(
                [
                    monitor.name,
                    getattr(
                        monitor,
                        "status",
                        "",
                    ),
                    uptime,
                    response_data["average_ms"],
                    summary_data["checks"],
                    summary_data["successful_checks"],
                    summary_data["failed_checks"],
                    summary_data["incidents"],
                    format_duration(summary_data["downtime_seconds"]),
                ]
            )

            summary_row += 1

            # Status
            style_status_cell(
                summary.cell(
                    row=row,
                    column=2,
                ),
                getattr(
                    monitor,
                    "status",
                    "",
                ),
            )

            # Bordi
            for column in range(
                1,
                10,
            ):
                summary.cell(
                    row=row,
                    column=column,
                ).border = THIN_BORDER

        # --------------------------------------------------
        # TOTALE / MEDIA
        # --------------------------------------------------

        if uptime_values:

            average_uptime = round(
                sum(uptime_values) / len(uptime_values),
                2,
            )

        else:

            average_uptime = None

        total_row = summary.max_row + 1

        summary.append(
            [
                "Totale / Media",
                "",
                average_uptime,
                selected_response_average,
                checks_total,
                successful_total,
                failed_total,
                incidents_total,
                format_duration(downtime_total),
            ]
        )

        style_total_row(
            summary,
            total_row,
            1,
            9,
        )

        # --------------------------------------------------
        # FORMATI NUMERICI
        # --------------------------------------------------

        for row in range(
            5,
            total_row + 1,
        ):

            summary.cell(
                row=row,
                column=3,
            ).number_format = '0.00"%"'

            summary.cell(
                row=row,
                column=4,
            ).number_format = '0.00 "ms"'

            for column in (
                5,
                6,
                7,
                8,
            ):
                summary.cell(
                    row=row,
                    column=column,
                ).number_format = "#,##0"

            summary.cell(
                row=row,
                column=9,
            ).number_format = "[h]:mm:ss"

        # --------------------------------------------------
        # FILTRO + FREEZE
        # --------------------------------------------------

        summary.auto_filter.ref = f"A4:I{total_row - 1}"

        summary.freeze_panes = "A5"

        # --------------------------------------------------
        # LARGHEZZE
        # --------------------------------------------------

        widths = {
            "A": 32,
            "B": 14,
            "C": 14,
            "D": 22,
            "E": 13,
            "F": 13,
            "G": 13,
            "H": 13,
            "I": 18,
        }

        for column, width in widths.items():

            summary.column_dimensions[column].width = width

        # Allineamento
        for row in summary.iter_rows(
            min_row=5,
            max_row=total_row,
            min_col=3,
            max_col=9,
        ):

            for cell in row:
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                )

    # ==================================================
    # SCHEDA SINGOLO MONITOR
    # ==================================================

    for item in monitor_data:

        monitor = item["monitor"]

        statistics = item["statistics"]

        sheet_name = make_unique_sheet_name(
            workbook,
            monitor,
        )

        sheet = workbook.create_sheet(sheet_name)

        summary_data = statistics["summary"]

        response_summary = statistics["response_time"]

        # --------------------------------------------------
        # TITOLO
        # --------------------------------------------------

        sheet.merge_cells("A1:H1")

        sheet["A1"] = monitor.name

        sheet["A1"].fill = PatternFill(
            fill_type="solid",
            fgColor=TITLE_FILL,
        )

        sheet["A1"].font = Font(
            bold=True,
            size=16,
            color=TITLE_FONT,
        )

        sheet["A1"].alignment = Alignment(
            vertical="center",
        )

        sheet.row_dimensions[1].height = 28

        # --------------------------------------------------
        # PERIODO
        # --------------------------------------------------

        sheet.merge_cells("A2:H2")

        sheet["A2"] = (
            f"Periodo: {period}   |   "
            f"Generato il: "
            f"{timezone.localtime(now):%d/%m/%Y %H:%M}"
        )

        sheet["A2"].font = Font(
            color="495057",
            italic=True,
        )

        # --------------------------------------------------
        # STATISTICHE RIASSUNTIVE
        # --------------------------------------------------

        kpi_labels = [
            "Uptime",
            "Response medio",
            "Response minimo",
            "Response massimo",
            "Checks",
            "Checks OK",
            "Checks KO",
            "Incidenti",
        ]

        kpi_values = [
            summary_data["uptime_percentage"],
            response_summary["average_ms"],
            response_summary["min_ms"],
            response_summary["max_ms"],
            summary_data["checks"],
            summary_data["successful_checks"],
            summary_data["failed_checks"],
            summary_data["incidents"],
        ]

        # Prima riga KPI
        for index in range(8):

            column = index + 1

            label_cell = sheet.cell(
                row=4,
                column=column,
            )

            value_cell = sheet.cell(
                row=5,
                column=column,
            )

            label_cell.value = kpi_labels[index]

            label_cell.fill = PatternFill(
                fill_type="solid",
                fgColor=SUBTITLE_FILL,
            )

            label_cell.font = Font(
                bold=True,
                color=SUBTITLE_FONT,
            )

            label_cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

            value_cell.value = kpi_values[index]

            value_cell.font = Font(
                bold=True,
                size=11,
            )

            value_cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

            value_cell.border = THIN_BORDER

        sheet.row_dimensions[4].height = 20
        sheet.row_dimensions[5].height = 24

        # Formati KPI
        sheet["A5"].number_format = '0.00"%"'

        for cell_reference in (
            "B5",
            "C5",
            "D5",
        ):
            sheet[cell_reference].number_format = '0.00 "ms"'

        for cell_reference in (
            "E5",
            "F5",
            "G5",
            "H5",
        ):
            sheet[cell_reference].number_format = "#,##0"

        # --------------------------------------------------
        # DOWNTIME
        # --------------------------------------------------

        sheet.merge_cells("A7:B7")

        sheet["A7"] = "Downtime periodo"

        sheet["A7"].fill = PatternFill(
            fill_type="solid",
            fgColor=SUBTITLE_FILL,
        )

        sheet["A7"].font = Font(
            bold=True,
            color=SUBTITLE_FONT,
        )

        sheet["A7"].alignment = Alignment(
            horizontal="left",
            vertical="center",
        )

        sheet.merge_cells("C7:D7")

        sheet["C7"] = format_duration(summary_data["downtime_seconds"]
)
        sheet["C7"].number_format = "[h]:mm:ss"

        sheet["C7"].font = Font(
            bold=True,
        )

        sheet.merge_cells("E7:H7")

        sheet["E7"] = "Dati temporali"

        sheet["E7"].font = Font(
            italic=True,
            color="6C757D",
        )

        # --------------------------------------------------
        # TABELLA TEMPORALE
        # --------------------------------------------------

        headers = [
            "Data/ora",
            "Uptime",
            "Response medio (ms)",
            "Checks totali",
            "Checks OK",
            "Checks KO",
            "Incidenti",
            "Downtime",
        ]

        for column, header in enumerate(
            headers,
            start=1,
        ):

            sheet.cell(
                row=9,
                column=column,
            ).value = header

        style_header(
            sheet,
            9,
        )

        # --------------------------------------------------
        # DATI
        # --------------------------------------------------

        response_data = {
            item["date"]: item["average_ms"]
            for item in statistics["response_time_over_time"]
        }

        checks_data = {item["date"]: item for item in statistics["checks"]}

        incidents_data = {item["date"]: item for item in statistics["incidents"]}

        uptime_data = {item["date"]: item for item in statistics["uptime"]}

        # PIÙ RECENTE -> PIÙ VECCHIO
        dates = sorted(
            uptime_data.keys(),
            reverse=True,
        )

        for date in dates:

            uptime_item = uptime_data[date]

            check_item = checks_data.get(
                date,
                {
                    "successful": 0,
                    "failed": 0,
                },
            )

            incident_item = incidents_data.get(
                date,
                {
                    "count": 0,
                    "downtime_seconds": 0,
                },
            )

            row = sheet.max_row + 1

            total_checks = check_item["successful"] + check_item["failed"]

            sheet.append(
                [
                    format_datetime(date),
                    uptime_item["uptime_percentage"],
                    response_data.get(date),
                    total_checks,
                    check_item["successful"],
                    check_item["failed"],
                    incident_item["count"],
                    format_duration(incident_item["downtime_seconds"]),
                ]
            )

            sheet.cell(
                row=row,
                column=1,
            ).number_format = "DD/MM/YYYY HH:MM"

            sheet.cell(
                row=row,
                column=2,
            ).number_format = '0.00"%"'

            sheet.cell(
                row=row,
                column=3,
            ).number_format = '0.00 "ms"'

            for column in (
                4,
                5,
                6,
                7,
            ):
                sheet.cell(
                    row=row,
                    column=column,
                ).number_format = "#,##0"

            sheet.cell(
                row=row,
                column=8,
            ).number_format = "[h]:mm:ss"

            for column in range(
                1,
                9,
            ):
                sheet.cell(
                    row=row,
                    column=column,
                ).border = THIN_BORDER

        # --------------------------------------------------
        # FILTRO + FREEZE
        # --------------------------------------------------

        last_row = sheet.max_row

        sheet.auto_filter.ref = f"A9:H{last_row}"

        sheet.freeze_panes = "A10"

        # --------------------------------------------------
        # LARGHEZZE
        # --------------------------------------------------

        widths = {
            "A": 20,
            "B": 14,
            "C": 22,
            "D": 15,
            "E": 13,
            "F": 13,
            "G": 13,
            "H": 18,
        }

        for column, width in widths.items():

            sheet.column_dimensions[column].width = width

        # --------------------------------------------------
        # ALLINEAMENTO
        # --------------------------------------------------

        for row in sheet.iter_rows(
            min_row=10,
            max_row=last_row,
            min_col=2,
            max_col=8,
        ):

            for cell in row:

                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                )

    return workbook


# ======================================================
# EXPORT HTTP
# ======================================================


def workbook_to_file_response(workbook):

    output = BytesIO()

    workbook.save(output)

    output.seek(0)

    return output.read()
